from flask import Flask
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from create import table1, table2, table3

# 数据库的配置变量
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:1q,4/ZzKkLl@127.0.0.1:3306/爬虫实验1'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = True
app.config['SQLALCHEMY_ECHO'] = True
db = SQLAlchemy(app)


# 批量获取表格数据
def get_data():

    # 打开 Excel 表格并获取表格名称
    filename1 = r'C:\Users\87898\Desktop\Linux\SRP项目内容\数据库相关内容\最终版\SRP-data-part\数据录入版.xlsx'    # 输入要读取的文档路径
    workbook = load_workbook(filename=filename1)

    # 获取表格
    Sheet1 = workbook['实验室表']

    # 获取文件内容
    file_content = []
    temp = []
    for i in Sheet1.iter_rows(min_row=311, max_row=322, min_col=1, max_col=5):
        for j in i:
            temp.append(j.value)
        file_content.append(temp)
        temp = []

    # 批量插入数据到数据库中
    # 学院与楼栋表：record = table1(building_id=file_content[i][0], building=file_content[i][1], faculty=file_content[i][2],
    # total_floors=file_content[i][3])
    # 中间表：record = table2(location_id=file_content[i][0], specific_floor=file_content[i][1],
    # building_id=file_content[i][2])
    # 实验室表：record = table3(laboratory_id=file_content[i][0], laboratory_location=file_content[i][1],
    # laboratory_name=file_content[i][2], remarks=file_content[i][3], location_id=file_content[i][4])
    for i in range(len(file_content)):
        record = table3(laboratory_id=file_content[i][0], laboratory_location=file_content[i][1],
                        laboratory_name=file_content[i][2], remarks=file_content[i][3], location_id=file_content[i][4])
        print(record)
        db.session.add(record)
        print('数据插入成功')
        db.session.commit()


if __name__ == '__main__':
    # 批量数据的插入
    get_data()

    # 单条数据插入
    '''
    record1 = table1(building_id=1, building='B1a', faculty='微电子学院', total_floors=9)
    db.session.add(record1)
    db.session.commit()
    '''
