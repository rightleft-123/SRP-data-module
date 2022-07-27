# coding=utf-8
import pymysql
from openpyxl import load_workbook


# 获取表格数据
def get_data():

    # 1.打开 Excel 表格并获取表格名称
    filename1 = ''    # 修改1：文档路径
    workbook = load_workbook(filename=filename1)

    # 2.获取表格
    Sheet1 = workbook['实验室表']

    file_content = []
    temp = []
    # 获取文件内容
    for i in Sheet1.iter_rows(min_row=2, max_row=321, min_col=1, max_col=5):
        for j in i:
            temp.append(j.value)
        file_content.append(temp)
        temp = []
    print(file_content)

    #插入数据到数据库中
    for i in range(len(file_content)):
        CON(file_content[i][0], file_content[i][1], file_content[i][2], file_content[i][3], file_content[i][4])


# 定义创建数据库的函数
def create():
    db = pymysql.connect(host='localhost',
                         user='root',
                         password='',
                         database='')    # 修改2：数据库信息

    # 使用 cursor() 方法创建一个游标对象 cursor
    print(1)
    cursor = db.cursor()

    # 使用 execute() 方法执行 SQL，如果表存在则删除
    print(1)
    cursor.execute("DROP TABLE IF EXISTS 实验室表")

    # 使用预处理语句创建表
    print(1)
    sql = """CREATE TABLE 实验室表 (
            实验室id int primary key,
            实验室位置 VARCHAR(10),
            实验室名称 VARCHAR(25),
            位置id int,
            备注 VARCHAR(100),
            foreign key (位置id)REFERENCES 中间表(位置id)
            )engine=InnoDB,charset=utf8mb4
            """
    cursor.execute(sql)
    # 关闭数据库
    print(1)
    db.close


# 定义一个连接数据库的函数value1, value2, value3
def CON(value1, value2, value3, value4, value5):
    db = pymysql.connect(host='localhost',
                         user='root',
                         password='',
                         database=''
                         )

    # 使用 cursor() 方法创建一个游标对象 cursor
    cursor = db.cursor()
    sql = f'insert into 实验室表 (实验室id,实验室位置,实验室名称,位置id,备注) values({value1},"{value2}","{value3}",{value4},"{value5}")'


    try:
        # 执行sql语句
        cursor.execute(sql)
        # 执行sql语句
        db.commit()
        print('插入数据成功')
    except:
        # 发生错误时回滚
        db.rollback()
        print('插入数据失败')
    # 关闭数据库连接
    db.close()


# create()
# CON()
# get_data()
